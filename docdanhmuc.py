#!/usr/bin/python3
import defusedxml
from defusedxml.common import EntitiesForbidden
import openpyxl


defusedxml.defuse_stdlib()

def secure_open_file(tenfile):
    try:
        wb = openpyxl.load_workbook(tenfile)
        return wb
    except EntitiesForbidden:
        raise ValueError('Vui lòng sử dụng tệp xlsx không có XEE')

def docnoidung(workbook, tensheet):
    noidung_sheet = workbook.get_sheet_by_name(tensheet)
    mang_tencot = []
    so_hang = noidung_sheet.max_row
    so_cot = noidung_sheet.max_column
    return so_hang, so_cot

if __name__ == '__main__':
    file = './danhmuc.xlsx'
    mofile = secure_open_file(file)
    danhsach_sheet = mofile.get_sheet_names()
    for i in danhsach_sheet:
        hang, cot = docnoidung(mofile, i)
        print(hang)
        print(cot)

